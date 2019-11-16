import openpyxl
from openpyxl import Workbook, load_workbook
import pymongo

myclient = pymongo.MongoClient("mongodb://localhost:27017/")
mydb = myclient["fashion"]
source = mydb["colormodels"]
target = mydb["dressmaininfos"]


path = "./apparels_maininfo.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

inserter = {}

p = 0

count = 0
for x in source.find():

    # print(x["page_color_link"])

    # print(count)

    # commented code

    buffer = []

    for dress in range(2, (sheet_obj.max_row)+1):
        m = sheet_obj.cell(row=dress, column=1)
        # print(m.value)

        inserter = {}

        if x["page_color_link"] == m.value:
            if m.value not in buffer:
                buffer.append(m.value)
                count = count+1
                print(count)

                try:

                    inserter["colorId"] = x["_id"]
                    # print(type(x["_id"]))
                    inserter["smallimageset"] = eval(
                        sheet_obj.cell(row=dress, column=2).value)
                    inserter["hdimageset"] = eval(
                        sheet_obj.cell(row=dress, column=3).value)
                    inserter["price"] = sheet_obj.cell(
                        row=dress, column=4).value
                    c = 0
                    try:
                        c = 1
                        k = int(sheet_obj.cell(row=dress, column=5).value)
                    except:
                        c = 2

                    if c == 1:
                        inserter["discount"] = int(
                            int(sheet_obj.cell(row=dress, column=5).value))
                        inserter["tag"] = str(
                            sheet_obj.cell(row=dress, column=6).value)
                        inserter["color_text"] = str(
                            sheet_obj.cell(row=dress, column=7).value)
                        inserter["description"] = eval(
                            sheet_obj.cell(row=dress, column=8).value)
                        inserter["size_n_fit"] = eval(
                            sheet_obj.cell(row=dress, column=9).value)

                    if c == 2:
                        inserter["discount"] = 0
                        inserter["tag"] = "notag"
                        inserter["color_text"] = str(
                            sheet_obj.cell(row=dress, column=5).value)
                        inserter["description"] = eval(
                            sheet_obj.cell(row=dress, column=6).value)
                        inserter["size_n_fit"] = eval(
                            sheet_obj.cell(row=dress, column=7).value)

                    p += 1

                    print(p)

                    target.insert_one(inserter)

                except:
                    print(str(m.value))

                # print(inserter)
